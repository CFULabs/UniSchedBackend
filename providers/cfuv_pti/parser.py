# Copyright (c) 2024, Kemran @remr2005
# Copyright (c) 2024, Alexander Baransky <alexander.baranskiy@yandex.ru>

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from base.types import LessonType, Lesson
import logging
from io import BytesIO

logger = logging.getLogger("cfuv_pti_parser")


DAYS_OF_WEEK = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота']
CYRILLIC_UPPER_LETTERS = 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ'
LESSON_TYPE_MAPPING = {
    "ПЗ": LessonType.practical,
    "ЛК": LessonType.lecture,
    "ЛР": LessonType.laboratory
}


def parse_xlsx(data: BytesIO) -> dict:
    """
    Функция, которая парсит .xlsx файл и возвращает словарь\n
    dict[группа][четная/нечетная неделя][день][номер пары][первая подгруппа или вторая(0/1)]\n
    сама пара, это тоже словарь с ключами type, name, teacher, location
    """
    schedule = {}
    reader = openpyxl.load_workbook(data)  # считываю файл
    for sheet_name in reader.sheetnames:
        sheet = reader[sheet_name]  # открываю нужный лист
        odd, even = None, None  # переменные для хранения начала четной и нечетной недели
        # находим координаты четной и нечетной недели
        row = 1
        for col in range(1, sheet.max_column):
            if odd and even:
                break

            val = str(sheet.cell(row, col).value).lower().strip()
            if val.startswith("нечетная"):
                odd = (row, col)
                continue

            if val.startswith("четная"):
                even = (row, col)
                continue

        if odd is None or even is None:
            logger.error(f"Couldn't find odd and even weeks coordinates in sheet {sheet}!")
            continue

        if odd[1] < even[1]:
            odd_res = parse_week(sheet, odd, even[1])
            even_res = parse_week(sheet, even, sheet.max_column)
        else:
            odd_res = parse_week(sheet, odd, sheet.max_column)
            even_res = parse_week(sheet, even, odd[1])

        for group in odd_res.keys():
            schedule[group] = {}

        for group, data in odd_res.items():
            schedule[group]["odd"] = data

        for group, data in even_res.items():
            schedule[group]["even"] = data

    return schedule


def is_cell_merged(sheet: Worksheet, cell: Cell) -> bool:
    for merged_range in sheet.merged_cells.ranges:  # Проверка, принадлежит ли ячейка к объединенным
        if cell.coordinate in merged_range:
            return True

    return False


def parse_lesson(sheet: Worksheet, row: int, col: int) -> dict[str, str]:
    name: str = sheet.cell(row, col).value
    assert name

    # КОСТЫЛИ alert!
    teacher: str = sheet.cell(row + 1, col).value
    if not teacher:
        # Try the next row
        row += 1
        teacher = sheet.cell(row + 1, col).value
    assert teacher
    # Trim prefix
    start_idx = 0
    for i in range(len(teacher)):
        if teacher[i] in CYRILLIC_UPPER_LETTERS:
            start_idx = i
            break
    teacher = teacher[start_idx:]

    location: str = sheet.cell(row + 2, col).value
    if not location:
        # Try the next row
        row += 1
        location = sheet.cell(row + 2, col).value
    assert location
    location = location[:location.find("(")].removeprefix("ауд.")

    return {
        "name": name,
        "teacher": teacher,
        "location": location
    }


def parse_week(sheet: Worksheet, coord: tuple[int, int], coord_end: int) -> dict:
    """
    Парсит лист эксель по определенной недели(четной/нечетной), возвращает \n
    dict, который составлен dict[группа][день][пара][группа(первая - 0, вторая - 1)] \n
    """
    schedule: dict[str, list] = {}
    cur_day: int | None = None
    groups = {}  # словарь для групп
    for row in range(1, 4):  # находим все группы для заданной недели
        for col in range(1, coord_end):
            cur_val = str(sheet.cell(row, col).value)
            if "группа" in cur_val:
                name = cur_val.removeprefix("группа").strip()
                groups[name] = (row, col)
                schedule[name + '(1)'] = [None] * len(DAYS_OF_WEEK)
                schedule[name + '(2)'] = [None] * len(DAYS_OF_WEEK)

    for row in range(5, sheet.max_row):
        dow = sheet.cell(row, coord[1]).value
        if dow and dow.lower() in DAYS_OF_WEEK:
            cur_day = DAYS_OF_WEEK.index(dow.lower())

        for group, (_, col) in groups.items():
            if not (lesson_type := sheet.cell(row, col - 1).value):
                continue

            lesson_num = int(sheet.cell(row, coord[1] + 1).value)
            first_cell, second_cell = sheet.cell(row, col), sheet.cell(row, col + 1)
            if not first_cell.value and not second_cell.value:
                continue

            def save_lesson(i: int, data: dict[str, str]):
                if not schedule[f"{group}({i})"][cur_day]:
                    schedule[f"{group}({i})"][cur_day] = []

                while len(schedule[f"{group}({i})"][cur_day]) < lesson_num - 1:
                    schedule[f"{group}({i})"][cur_day].append(None)

                Lesson.model_validate(data)  # Validate with pydantic
                schedule[f"{group}({i})"][cur_day].append(data)

            types = lesson_type.split("/")  # Maybe two types in one cell? (Even if only one lesson exists)
            try:
                if first_cell.value and second_cell.value:
                    # Two different lessons for both subgroups
                    save_lesson(1, {
                        "type": LESSON_TYPE_MAPPING.get(types[0]),
                        **parse_lesson(sheet, row, col)
                    })
                    save_lesson(2, {
                        "type": LESSON_TYPE_MAPPING.get(types[1] if len(types) > 1 else types[0]),
                        **parse_lesson(sheet, row, col + 1)
                    })
                else:
                    if first_cell.value:
                        # Lesson only for the first subgroup / same lesson for both subgroups
                        lesson_data = {
                            "type": LESSON_TYPE_MAPPING.get(types[0]),
                            **parse_lesson(sheet, row, col)
                        }
                        save_lesson(1, lesson_data)

                        # Check if cell is merged
                        if is_cell_merged(sheet, first_cell):
                            # Copy the same to the second subgroup
                            save_lesson(2, lesson_data)
                    else:
                        # Lesson only for the second subgroup
                        save_lesson(2, {
                            "type": LESSON_TYPE_MAPPING.get(types[0]),
                            **parse_lesson(sheet, row, col + 1)
                        })
            except (AssertionError, ValidationError) as e:
                logger.error(f"Failed to parse lesson at {first_cell.coordinate}-{second_cell.coordinate} in {sheet}!")
                logger.exception(e)

    return schedule
